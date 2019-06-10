import os
import sys
from os import listdir
from difflib import unified_diff
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from dd_folder_obj import DD_Folder
from dd_folder_obj import get_dd_file
from dd_folder_obj import fbf_diffs
import win32com.client
import datetime
import shutil

report_template = os.path.abspath("report_template.xlsm")
report = os.path.abspath("last_report.xlsm")

""" Description:

    Flags:
        -o  <FILENAME>  : Specifies the report name
        -n              : Specifies no Excel output
        -h              : Displays Help
"""
def main():

    #Strips dd-comp.py call from sys.argv
    args = sys.argv[1:]
    paths = sys.argv[1:]

    print("\n\nSTARTING:  ---------- dd-comp.py ----------")
    print("* sys.argv:\n    ->"+str(sys.argv))
    print("* args:\n    ->"+str(args))

    #Basic error checking
    call_error_checking(paths)
    print("* paths:\n    ->"+str(paths))

    #out_file by default should go to the Downloads folder with the time and date in the file name
    call_time = datetime.datetime.now().strftime("%m-%d-%Y_%Hh-%Mm")
    out_file = os.path.expanduser("~\\Desktop\\")+"report"+call_time+".xlsm"

    #Handle any necessary flags and find the two arguments
    if args[0] == "-o":
        out_file = args[1]
    elif args[0] == "-n":
        out_file = None

    #Handle paths and remove ending / for clarity
    path_0 = paths[0]
    if path_0[len(path_0) - 1] == '/':
        path_0 = path_0[:len(path_0) - 1]

    path_1 = paths[1]
    if path_1[len(path_1) - 1] == '/':
        path_1 = path_1[:len(path_1) - 1]

    #Create DD Folder objects and print their contents
    obj_0 = DD_Folder(path_0)
    print("\n\nINSIDE:  ---------- OBJ_0 ----------")
    print(str(obj_0))

    obj_1 = DD_Folder(path_1)
    print("\n\nINSIDE:  ---------- OBJ_1 ----------")
    print(str(obj_1))

    #Scrape the DD_Folders.  Append the Differeneces returned
    #Returns a list of differences in the form (Key, expected_value, file_index, actual_value)

    #Populate File-by-File Differences.
    #import dd_folder_obj import fdf_diffs
    value_diffs = fbf_diffs(obj_0, obj_1)

    #If out_file == None: pretty print differences
    if out_file == None: return

    #Else write to excel report
    write_xls(obj_0, obj_1, value_diffs, out_file)

    print("\n\nENDING:    ---------- dd-comp.py ----------")



""" Description: This function checks for usage errors.

    Minor Details:
        - Args expects the sys.argv with the call to
            the dd-comp.py file chopped off. Potentially
            makes expanding the number of ways to call
            easier.

    Current Implimentations:
        1. Checks validity of flags.
            -> For more information on implimented flags
                check header comment for main function.
        2. Compares '# of Args' to 'Expected Number of Args'.
        3. Checks validity of paths givin.
"""
def call_error_checking(args):

    #Invalid arguments, too few arguments
    if len(args) == 0:
        print("### - ERROR: Invalid number of arguments\n\t->Expected 2, recieved: 0")
        display_help()
        exit(1)

    #Check for 'HELP FLAG'
    if args[0] == "-h":
        display_help()
        exit(0)

    #Invalid usage flags
    if args[0] == "-o" or args[0] == "-n":
        if args[0] == "-o":
            #Invalid usage of '-o' flag
            if len(args) != 4:
                print("### - ERROR: Invalid number of arguments\n\t->Expected 4, recieved "+str(len(args)))
                display_help()
                exit(1)
            args.pop(1) #removes filename from arguments
        args.pop(0) #removes flag from arguments
    #Invalid flag
    elif len(args[0]) == 0 or args[0][0] == '-':
        print("### - ERROR: Invalid flag\n\t-> \""+str(args[0])+"\"")
        display_help()
        exit(1)

    #Invalid argument number
    if len(args) != 2:
        print("### - ERROR: Invalid number of arguments\n\t->Expected 2, recieved: "+str(len(args)))
        display_help()
        exit(1)

    #Invalid path
    Path_0 = args[0]
    Path_1 = args[1]
    if not os.path.exists(Path_0):
        print("### - ERROR: First path givin is invalid\n\t->Path 1: "+str(Path_0))
        display_help()
        exit(1)
    if not os.path.exists(Path_1):
        print("### - ERROR: Second path givin is invalid\n\t->Path 2: "+str(Path_1))
        display_help()
        exit(1)

    #Empty path
    if len(listdir(Path_0)) == 0:
        print("### - ERROR: First path givin is empty\n\t->Path 1: "+str(Path_0))
        exit(1)
    if len(listdir(Path_1)) == 0:
        print("### - ERROR: Second path givin is empty\n\t->Path 2: "+str(Path_1))
        exit(1)



""" Description: Displays help message in shell.
"""
def display_help():
    print("Usage: \'python dd-comp [-o <OUTFILE>|n|h] \"<PATH-1>\" \"<PATH-2>\"\'")
    print("\tFlags:")
    print("\t\t-o <OUTFILE> : Specifies the out_file")
    print("\t\t-n           : Specifies no Excel output")
    print("\t\t-h           : Displays help")



""" Description: Writes differences to the excel document.
"""
def write_xls(obj_0, obj_1, fbf, out_file):
    #Write First Page
    print("\n\n***STARTING:    ---------- write_xls(obj_0, obj_1, out_file) ----------")
    print("  * obj_0.folder_path: "+obj_0.folder_path)
    print("  * obj_1.folder_path: "+obj_1.folder_path)
    print("  * out_file: "+out_file)

    # Load excel document with macros to be run
    wb = load_workbook(report_template, read_only=False, keep_vba=True)
    wb.remove_sheet(wb["Sheet1"])

    """ Write fbf
        # Calls the file by file differences to be written
        # wb.create_sheet("alert fbf")
        # ws = wb["alert fbf"]
        # ws.column_dimensions[get_column_letter(1)].width = 45
        # write_fbf(ws, fbf)
    """

    #Write line-line Page in For loop
    print("  ***STARTING: line-line Loop")
    #obj_0 and obj_1 files indices
    i0, i1 = 0, 0
    obj_len0, obj_len1 = len(obj_0.files), len(obj_1.files)
    file_file_max = max(obj_len0, obj_len1)
    pass_three = []
    for i in range(file_file_max):
        #Check files indices
        if i0 >= obj_len0 or i1 >= obj_len1:
            break

        #If files are not the same files types or unexpected types
        file_ref0 = obj_0.file_info[i0][0]
        file_ref1 = obj_1.file_info[i1][0]
        if file_ref0 != file_ref1 or file_ref0 == -1:
            if file_ref0 > file_ref1:
                i1 += 1
            elif file_ref1 > file_ref0:
                i0 += 1
            else:
                i0 += 1
                i1 += 1
            continue

        #If the file is not a comparable file
        file_ref_name = obj_0.expected_files[file_ref0][0]
        comparable_file = obj_0.expected_files[file_ref0][1]
        if not comparable_file:
            i0 += 1
            i1 += 1
            continue

        #Move to or create excel sheet for file type
        sheet_name = str(file_ref_name)+" file"
        print("    *** "+sheet_name+" Modification")
        if sheet_name not in wb.sheetnames:
            create_file_sheet(wb, sheet_name)
        #ws = wb[sheet_name]

        # Compute and write the Diff (line-by-line) results
        path_0 = obj_0.folder_path+"/"+obj_0.files[i0]
        path_1 = obj_1.folder_path+"/"+obj_1.files[i1]
        fileType = obj_0.file_info[i0][0]
        print("      * path_0: "+path_0)
        print("      * path_1: "+path_1)
        pass_three.extend(diff_file(path_0, path_1, sheet_name, wb, file_type=fileType))

        #Save File
        wb.save(report)

        #increment files indices
        i0 += 1
        i1 += 1

    #Passthrough 3: Call Excel macro to color code the lines
    xl=win32com.client.Dispatch("Excel.Application")
    wb=xl.Workbooks.Open(report)
    file_name = report.split('\\')
    file_name = file_name[len(file_name)-1]
    try:
        for i in pass_three:
            xl.Application.Run(file_name+"!Module1.Highlight_Range", i[0],i[1],i[2],i[3],i[4])
    finally:
        wb.Close(SaveChanges=1)
    xl.Quit()

    shutil.copyfile(report, out_file)
    print("***ENDING:    ---------- write_xls(obj_0, obj_1, out_file) ----------")



""" Description: Creates a new sheet in an excel document with a specific sheet name.
"""
def create_file_sheet(wb, sheet_name):
    wb.create_sheet(sheet_name)
    ws = wb[sheet_name]

    #Format Sheet
    line_col_0 = 1
    item_col_0 = 2
    item_col_1 = 3
    line_col_1 = 4

    #Format Widths
    ws.column_dimensions[get_column_letter(item_col_0)].width = 60
    ws.column_dimensions[get_column_letter(item_col_1)].width = 60
    ws.column_dimensions[get_column_letter(item_col_0)].alignment = Alignment(wrap_text=True)
    ws.column_dimensions[get_column_letter(item_col_1)].alignment = Alignment(wrap_text=True)

    ws.column_dimensions[get_column_letter(line_col_0)].width = 5
    ws.column_dimensions[get_column_letter(line_col_1)].width = 5




""" Description: Writes differences to the excel document.
"""
def diff_file(path_0, path_1, sheet_name, wb, row_number=1, file_type=-1):
    print("      ***STARTING:     ---------- diff_file(path_0, path_1, ws) ----------")
    #Get worksheet
    ws = wb[sheet_name]
    to_highlight = []

    #Transcribe Old File
    file_0 = get_dd_file(path_0, file_type=file_type)

    #Transcribe New File
    file_1 = get_dd_file(path_1, file_type=file_type)

    #Getting the file diffs
    diff = h_diff(file_0, file_1)

    #Placing the proper values in the proper cells
    line_num_0 = 1
    contents_0 = 2
    contents_1 = 3
    line_num_1 = 4
    for line in diff:
        #Populate the line nubers
        if line[0] != None: ws.cell(row=row_number, column=line_num_0).value = line[0]
        if line[1] != None: ws.cell(row=row_number, column=line_num_1).value = line[1]

        #Populate the contents portion of file 0
        if line[2] != None:
            ws.cell(row=row_number, column=contents_0).value = line[2]
            #Populate to_highlight list for macro calling in the callee
            if line[6] != None:
                for x in line[6]:
                    to_highlight.append((row_number, contents_0, x[0], x[1], sheet_name))


        #Populate the contents portion of file 1
        if line[3] != None:
            ws.cell(row=row_number, column=contents_1).value = line[3]
            #Populate to_highlight list for macro calling in the callee
            if line[7] != None:
                for x in line[7]:
                    to_highlight.append((row_number, contents_1, x[0], x[1], sheet_name))


        #Format the cells - Has to be done because openpyxl clears formatting.
        ws.cell(row=row_number, column=line_num_0).fill = PatternFill(start_color=line[4], end_color=line[4], fill_type='solid')
        ws.cell(row=row_number, column=line_num_1).fill = PatternFill(start_color=line[5], end_color=line[5], fill_type='solid')
        ws.cell(row=row_number, column=contents_0).fill = PatternFill(start_color=line[4], end_color=line[4], fill_type='solid')
        ws.cell(row=row_number, column=contents_0).alignment = Alignment(wrap_text=True)
        ws.cell(row=row_number, column=contents_1).fill = PatternFill(start_color=line[5], end_color=line[5], fill_type='solid')
        ws.cell(row=row_number, column=contents_1).alignment = Alignment(wrap_text=True)

        row_number += 1

    return to_highlight

    """ Old Diff Style
        #Run Diff on Files
        diff = [line for line in unified_diff(file_0, file_1)]
        diff = delo(diff)
        print("        *** diff:")

        #Write diffs
        neg_line_col = 1
        neg_item_col = 2
        pos_item_col = 3
        pos_line_col = 4
        for i in diff:
            print("          * diff["+str(line_number)+"]: "+str(i))
            if i[1]:
                ws.cell(row=line_number, column=neg_line_col).value = i[0]
                ws.cell(row=line_number, column=neg_item_col).value = i[2]
            else:
                ws.cell(row=line_number, column=pos_line_col).value = i[0]
                ws.cell(row=line_number, column=pos_item_col).value = i[2]
            line_number += 1
    """



""" Description: Performs a diff that returns an array of lists with the information presented in the following manner
"""
def h_diff(file_0, file_1):
    #returned object
    h_list = []

    #Color Codes
    missing_grey = '8A8A8A'
    normal_white = 'FFFFFF'
    diff_red = 'F8C1A2'

    #Passthrough 1: Insert Matching lines and information and put in the proper ammount of lines
    i0, i1 = 0, 0
    #for i in the range(len(file_0)+len(file_1)):
    while i0 != len(file_0) and i1 != len(file_1):
        # print("i0: "+str(i0)+"| i1: "+str(i1))
        #find next occurence in file 0 or file 1
        if file_0[i0] == file_1[i1]:
            h_list.append((i0, i1, file_0[i0], file_1[i1], normal_white, normal_white, None, None))
            i0 += 1
            i1 += 1
        else:
            for diff_index in range(1, len(file_0) + len(file_1)):

                # print("diff_index: "+str(diff_index))
                if i0+diff_index >= len(file_0) and i1+diff_index >= len(file_1):
                    diff_marked_0, diff_marked_1, diff_ranges_0, diff_ranges_1 = l_diff(file_0[i0], file_1[i1])
                    h_list.append((i0, i1, diff_marked_0, diff_marked_1, diff_red, diff_red, diff_ranges_0, diff_ranges_1))
                    i0 += 1
                    i1 += 1
                    break
                elif i1+diff_index < len(file_1) and file_0[i0] == file_1[i1+diff_index]:
                    for catchup_i in range(diff_index):
                        h_list.append((None, i1+catchup_i, None, file_1[i1+catchup_i], missing_grey, diff_red, None, None))
                    i1 = i1+diff_index
                    break
                elif i0+diff_index < len(file_0) and file_0[i0+diff_index] == file_1[i1]:
                    for catchup_i in range(diff_index):
                        h_list.append((i0+catchup_i, None, file_0[i0+catchup_i], None, diff_red, missing_grey, None, None))
                    i0 = i0+diff_index
                    break
    if i0 != len(file_0):
        for i in range(i0, len(file_0)):
            h_list.append((i, None, file_0[i], None, diff_red, missing_grey, None, None))
    elif i1 != len(file_1):
        for i in range(i1, len(file_1)):
            h_list.append(( None, i, None, file_1[i], missing_grey, diff_red, None, None))
    return h_list

    #Passthrough 2: Insert Different lines and information replacing problem spaces with underscores



""" Description: Returns difference ranges and the lines the with different spaces shown.
"""
def l_diff(line_0, line_1):
    words_0 = list(filter(lambda x: x!='', line_0.split(' ')))
    words_1 = list(filter(lambda x: x!='', line_1.split(' ')))
    line_0 = list(line_0)
    line_1 = list(line_1)
    diff_ranges_0 = []
    diff_ranges_1 = []
    i0, i1 = 0, 0
    w0, w1 = 0, 0
    while i0 != len(line_0) and i1 != len(line_1):
        # print("i0: "+str(i0)+"| i1: "+str(i1)+"| w0: "+str(w0)+"| w1: "+str(w1)+"| diff_ranges_0: "+str(diff_ranges_0)+"| diff_ranges_1: "+str(diff_ranges_1))
        #if the characters are both spaces
        if line_0[i0] == ' ' and line_1[i1] == ' ':
            i0 += 1
            i1 += 1
        else:
            # Replace extra spaces and add to highlight list
            if line_0[i0] == ' ':
                diff_ranges_0.append((i0, i0+1))
                line_0[i0] = '<Space>'
                i0+=1
            elif line_1[i1] == ' ':
                diff_ranges_1.append((i1, i1+1))
                line_1[i1] = '<Space>'
                i1+=1

            #if neither of the characters are spaces
            else:
                #if the words are the same
                if words_0[w0] == words_1[w1]:
                    i0 += len(words_0[w0])
                    i1 += len(words_1[w1])
                    w0 += 1
                    w1 += 1
                #words aren't the same
                else:
                    last_words_are_different = True
                    for diff_index in range(1, len(words_0) + len(words_1)):
                        last_words_are_different = False
                        #catch non-matching words
                        if w0+diff_index >= len(words_0) and w1+diff_index >= len(words_1):
                            diff_ranges_0.append((i0, i0+len(words_0[w0])))
                            diff_ranges_1.append((i1, i1+len(words_1[w1])))
                            i0 += len(words_0[w0])
                            i1 += len(words_1[w1])
                            w0 += 1
                            w1 += 1
                            break

                        #found a matching words
                        elif w1+diff_index < len(words_1) and words_0[w0] == words_1[w1+diff_index]:
                            #calculate and append difference ranges for words_1
                            diff_i1 = i1
                            for catchup_i in range(diff_index):
                                diff_i1 += len(words_1[w1+catchup_i])
                                while diff_i1 < len(line_1) and line_1[diff_i1] == ' ': diff_i1 += 1

                            diff_ranges_1.append((i1, diff_i1))

                            w1 += diff_index
                            i1 = diff_i1
                            break

                        elif w0+diff_index < len(words_0) and words_0[w0+diff_index] == words_1[w1]:
                            #calculate and append difference ranges for words_0
                            diff_i0 = i0
                            for catchup_i in range(diff_index):
                                diff_i0 += len(words_0[w0+catchup_i])
                                while diff_i0 < len(line_0) and line_0[diff_i0] == ' ': diff_i0 += 1

                            diff_ranges_0.append((i0, diff_i0))

                            w0 += diff_index
                            i0 = diff_i0
                            break

                    if last_words_are_different:
                        diff_ranges_0.append((i0, i0+len(words_0[w0])))
                        diff_ranges_1.append((i1, i1+len(words_1[w1])))
                        i0 += len(words_0[w0])
                        i1 += len(words_1[w1])
    #if there is space left append it as final difference
    if i0 != len(line_0):
        diff_ranges_0.append((i0, len(line_0)))
        for i in range(i0, len(line_0)):
            if line_0[i] == ' ':
                line_0[i] = '_'
    elif i1 != len(line_1):
        diff_ranges_1.append((i1, len(line_1)))
        for i in range(i1, len(line_1)):
            if line_1[i] == ' ':
                line_1[i] = '_'

    #reduce ranges in diff_ranges_0 and diff_ranges_1
    initial_len=len(diff_ranges_0)
    for i, j in zip(range(initial_len-1,0,-1), range(initial_len-2,-1,-1)):
        if diff_ranges_0[j][1] == diff_ranges_0[i][0]:
            diff_ranges_0[j] = (diff_ranges_0[j][0], diff_ranges_0[i][1])
            diff_ranges_0.pop(i)
    initial_len=len(diff_ranges_1)
    for i, j in zip(range(initial_len-1,0,-1), range(initial_len-2,-1,-1)):
        if diff_ranges_1[j][1] == diff_ranges_1[i][0]:
            diff_ranges_1[j] = (diff_ranges_1[j][0], diff_ranges_1[i][1])
            diff_ranges_1.pop(i)

    #Returns a tuple
    return (''.join(line_0), ''.join(line_1), diff_ranges_0, diff_ranges_1)



""" Description: Writes differences to the excel document.
"""
def delo(lines):
  ln_a, ln_b = 1, 1
  diffs = []
  for i in range(3, len(lines)):
    polarity = lines[i][0]
    if polarity == '-':
      diff_to_add = (ln_a, True, lines[i][1:])
      ln_a += 1
    elif polarity == '+':
      diff_to_add = (ln_b, False, lines[i][1:])
      ln_b += 1
    else:
      ln_a += 1
      ln_b += 1
      continue
    diffs.append(diff_to_add)
  return diffs



""" Description: Write the fbf differences.
"""
def write_fbf(ws, fbf, spacing="", line_number=1):
    for i in fbf:
        if type(i) == list:
            if type(i[3]) == list or type(i[3]) == tuple:
                ws.cell(row=line_number, column=1).value = spacing+str(i[1:3])
                line_number += 1
                line_number = write_fbf(ws, i[3], spacing+"---", line_number=line_number)
            else:
                ws.cell(row=line_number, column=1).value = spacing+str(i[1:4])
                line_number += 1
                line_number = write_fbf(ws, i[4], spacing+"---", line_number=line_number)

        elif type(i) == tuple:
            if len(i) == 4 and (len(str(i[2])) > 15 or len(str(i[3])) > 15):
                #ws.cell(row=line_number, column=1).value = spacing+"("+str(i[0])+", "+str(i[1])+", "+str(i[2])[:15]+", "+str(i[3])[:15]+")"
                ws.cell(row=line_number, column=1).value = spacing+"("+str(i[1])+", "+str(i[2])[:15]+", "+str(i[3])[:15]+")"
                line_number += 1
            else:
                ws.cell(row=line_number, column=1).value = spacing+str(i[1:])
                line_number += 1
        elif len(str(i))>10:
            ws.cell(row=line_number, column=1).value = spacing+str(i)[:10]
            line_number += 1
        else:
            ws.cell(row=line_number, column=1).value = spacing+str(i)
            line_number += 1
    return line_number







""" Description: Calls the main funtion if this is the "main".
"""
if __name__ == "__main__": main()
